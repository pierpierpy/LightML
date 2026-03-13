import argparse
import json
import sqlite3
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.filters import AutoFilter


# ─────────────────────────────────────────────
# Color palette
# ─────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_OVERVIEW_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_OVERVIEW_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)

_SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SECTION_FONT = Font(bold=True, size=11, color="1F3864")

_THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)

_MODEL_FONT = Font(bold=True, size=10)
_CHECKPOINT_FONT = Font(italic=True, size=10, color="666666")


def _auto_width(ws, min_width=10, max_width=45, padding=3):
    """Set column widths based on the content of each column."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        best = min_width
        for cell in col_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length + padding > best:
                    best = length + padding
        ws.column_dimensions[col_letter].width = min(best, max_width)


# ─────────────────────────────────────────────
# Data loading
# ─────────────────────────────────────────────

def load_from_db(db_path: Path):

    rows = []
    metrics_map = {}
    families = {}
    runs = []
    models_info = []

    with sqlite3.connect(db_path) as conn:

        # RUNS
        run_rows = conn.execute("""
            SELECT id, run_name, description, metadata, created_at
            FROM run ORDER BY created_at
        """).fetchall()

        for rid, name, desc, meta, created in run_rows:
            runs.append({
                "id": rid,
                "run_name": name,
                "description": desc or "",
                "metadata": meta or "",
                "created_at": created or "",
            })

        # MODELS (with run info and parent)
        model_rows = conn.execute("""
            SELECT m.id, m.model_name, m.path, m.parent_id, m.notes, m.hidden,
                   r.run_name,
                   p.model_name AS parent_name
            FROM model m
            JOIN run r ON m.run_id = r.id
            LEFT JOIN model p ON m.parent_id = p.id
            ORDER BY r.run_name, m.model_name
        """).fetchall()

        for mid, name, path, parent_id, notes, hidden, run_name, parent_name in model_rows:
            models_info.append({
                "id": mid,
                "model_name": name,
                "path": path or "",
                "parent_name": parent_name or "",
                "notes": notes or "",
                "hidden": bool(hidden),
                "run_name": run_name,
            })
            rows.append({
                "key": ("model", mid),
                "name": name,
                "phase": "Model",
                "run_name": run_name,
                "parent": parent_name or "",
                "notes": notes or "",
            })

        # CHECKPOINTS
        ckpt_rows = conn.execute("""
            SELECT c.id, c.step, m.model_name, r.run_name
            FROM checkpoint c
            JOIN model m ON c.model_id = m.id
            JOIN run r ON m.run_id = r.id
            ORDER BY r.run_name, m.model_name, c.step
        """).fetchall()

        for cid, step, model_name, run_name in ckpt_rows:
            rows.append({
                "key": ("ckpt", cid),
                "name": f"{model_name} (step {step})",
                "phase": "Checkpoint",
                "run_name": run_name,
                "parent": model_name,
                "notes": "",
            })

        # METRICS
        metric_rows = conn.execute("""
            SELECT model_id, checkpoint_id, family, metric_name, value
            FROM metrics
        """).fetchall()

        for model_id, checkpoint_id, family, metric_name, value in metric_rows:
            if model_id:
                key = ("model", model_id)
            elif checkpoint_id:
                key = ("ckpt", checkpoint_id)
            else:
                continue

            if family not in families:
                families[family] = set()
            families[family].add(metric_name)

            if key not in metrics_map:
                metrics_map[key] = {}
            if family not in metrics_map[key]:
                metrics_map[key][family] = {}
            metrics_map[key][family][metric_name] = value

    return rows, metrics_map, families, runs, models_info


# ─────────────────────────────────────────────
# Overview sheet
# ─────────────────────────────────────────────

def _build_overview(wb, db_path, runs, models_info, families, rows):
    ws = wb.create_sheet(title="Overview", index=0)

    r = 1

    # ── Title ──
    ws.cell(row=r, column=1, value=f"LightML Report  -  {db_path.stem}")
    ws.cell(row=r, column=1).font = Font(bold=True, size=16, color="1F3864")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    ws.cell(row=r, column=1, value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=r, column=1).font = Font(size=10, color="808080")
    r += 2

    # ── Summary stats ──
    _section_row(ws, r, "Summary")
    r += 1
    n_models = sum(1 for ro in rows if ro["phase"] == "Model")
    n_ckpts = sum(1 for ro in rows if ro["phase"] == "Checkpoint")
    n_metrics = sum(len(v) for v in families.values())
    stats = [
        ("Runs", len(runs)),
        ("Models", n_models),
        ("Checkpoints", n_ckpts),
        ("Metric Families", len(families)),
        ("Total Metrics", n_metrics),
    ]
    for label, val in stats:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1

    # ── Runs ──
    if runs:
        _section_row(ws, r, "Runs")
        r += 1
        for hdr_idx, hdr in enumerate(["Run Name", "Description", "Metadata", "Created"], 1):
            c = ws.cell(row=r, column=hdr_idx, value=hdr)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = _HEADER_ALIGN
        r += 1
        for run in runs:
            ws.cell(row=r, column=1, value=run["run_name"])
            ws.cell(row=r, column=2, value=run["description"])
            meta = run["metadata"]
            if meta:
                try:
                    parsed = json.loads(meta)
                    meta = ", ".join(f"{k}: {v}" for k, v in parsed.items())
                except (json.JSONDecodeError, AttributeError):
                    pass
            ws.cell(row=r, column=3, value=meta)
            ws.cell(row=r, column=4, value=run["created_at"])
            r += 1
        r += 1

    # ── Models with notes ──
    if models_info:
        _section_row(ws, r, "Models")
        r += 1
        headers = ["Model", "Run", "Parent", "Path", "Notes", "Hidden"]
        for hdr_idx, hdr in enumerate(headers, 1):
            c = ws.cell(row=r, column=hdr_idx, value=hdr)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = _HEADER_ALIGN
        r += 1
        for m in models_info:
            ws.cell(row=r, column=1, value=m["model_name"]).font = Font(bold=True, size=10)
            ws.cell(row=r, column=2, value=m["run_name"])
            ws.cell(row=r, column=3, value=m["parent_name"])
            ws.cell(row=r, column=4, value=m["path"])
            ws.cell(row=r, column=5, value=m["notes"])
            ws.cell(row=r, column=6, value="Yes" if m["hidden"] else "")
            r += 1
        r += 1

    # ── Metric families ──
    if families:
        _section_row(ws, r, "Metric Families")
        r += 1
        for hdr_idx, hdr in enumerate(["Family", "Metrics"], 1):
            c = ws.cell(row=r, column=hdr_idx, value=hdr)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = _HEADER_ALIGN
        r += 1
        for family in sorted(families):
            ws.cell(row=r, column=1, value=family).font = Font(bold=True, size=10)
            ws.cell(row=r, column=2, value=", ".join(sorted(families[family])))
            r += 1

    _auto_width(ws, min_width=12, max_width=60)


def _section_row(ws, row, title):
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = _SECTION_FILL
    c = ws.cell(row=row, column=1, value=title)
    c.font = _SECTION_FONT
    c.fill = _SECTION_FILL


# ─────────────────────────────────────────────
# Metric sheets
# ─────────────────────────────────────────────

def _build_metric_sheet(wb, family, metric_names_set, rows, metrics_map):
    # Sanitize sheet title (Excel max 31 chars)
    title = family[:31]
    ws = wb.create_sheet(title=title)

    metric_names = sorted(metric_names_set)
    columns = ["Model", "Phase", "Run", "Parent"] + metric_names

    # ── Header ──
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # ── Data ──
    for row_idx, row in enumerate(rows, start=2):
        key = row["key"]
        family_metrics = metrics_map.get(key, {}).get(family, {})

        name_cell = ws.cell(row=row_idx, column=1, value=row["name"])
        phase_cell = ws.cell(row=row_idx, column=2, value=row["phase"])
        ws.cell(row=row_idx, column=3, value=row["run_name"])
        ws.cell(row=row_idx, column=4, value=row["parent"])

        # Style models vs checkpoints differently
        if row["phase"] == "Model":
            name_cell.font = _MODEL_FONT
        else:
            name_cell.font = _CHECKPOINT_FONT
            phase_cell.font = _CHECKPOINT_FONT

        for col_idx, metric in enumerate(metric_names, start=5):
            val = family_metrics.get(metric)
            if val is not None:
                ws.cell(row=row_idx, column=col_idx, value=val)

    max_row = ws.max_row

    # ── Conditional formatting on metric columns ──
    if max_row > 1:
        for col_idx in range(5, len(columns) + 1):
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{max_row}",
                ColorScaleRule(
                    start_type='min', start_color='F8696B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='63BE7B',
                ),
            )

    # ── AutoFilter (covers all columns) ──
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{max_row}"

    # Default filter: hide checkpoints (show only "Model" in Phase column)
    ws.auto_filter.add_filter_column(1, ["Model"])

    # ── Freeze header + model name column ──
    ws.freeze_panes = "C2"

    # ── Auto-width ──
    _auto_width(ws)


# ─────────────────────────────────────────────
# Main entry-point
# ─────────────────────────────────────────────

def export_excel(db_path: Path, output_path: Path):

    rows, metrics_map, families, runs, models_info = load_from_db(db_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Overview sheet (first)
    _build_overview(wb, db_path, runs, models_info, families, rows)

    # One sheet per metric family
    for family in sorted(families):
        _build_metric_sheet(wb, family, families[family], rows, metrics_map)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Export completed: {output_path}")


def main():

    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True)
    parser.add_argument("--output")

    args = parser.parse_args()

    db_path = Path(args.db)

    if args.output:
        output_path = Path(args.output)
    else:
        date = datetime.now().strftime("%Y-%m-%d")
        output_path = Path("report") / f"{db_path.stem}_report_{date}.xlsx"

    export_excel(db_path, output_path)


if __name__ == "__main__":
    main()
